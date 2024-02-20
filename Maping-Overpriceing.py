from fuzzywuzzy import process
import openpyxl

# Load the Winter Top Selling Excel file
winter_top_selling_file = "C:\\Users\\Snapp\\Desktop\\winter top selling (1).xlsx"
winter_top_selling_workbook = openpyxl.load_workbook(winter_top_selling_file)
winter_top_selling_sheet = winter_top_selling_workbook["top selling oct-dec"]

# Load the Overpricing Excel file
overpricing_file = "C:\\Users\\Snapp\\Desktop\\Overpricing.xlsx"
overpricing_workbook = openpyxl.load_workbook(overpricing_file)
overpricing_sheet = overpricing_workbook.active

# Create a new Excel workbook to store the mapped products
output_workbook = openpyxl.Workbook()
output_sheet = output_workbook.active

# Set headers for the output sheet
output_sheet['A1'] = 'Product Name'
output_sheet['B1'] = 'Product id'
output_sheet['C1'] = 'Price'

# Create a dictionary to store product names and their corresponding product IDs
product_id_map = {}

# Iterate through the Winter Top Selling file to extract product IDs
for row in winter_top_selling_sheet.iter_rows(min_row=2, values_only=True):
    product_name, product_id = row[0], row[1]
    product_id_map[product_name] = product_id

# Iterate through the Overpricing file to extract prices
row_index = 2  # Start writing from the second row
for row in overpricing_sheet.iter_rows(min_row=2, values_only=True):
    product_name, price = row[0], row[1]

    # Check if the product name exists in the Winter Top Selling file
    if product_name in product_id_map:
        product_id = product_id_map[product_name]
        output_sheet.cell(row=row_index, column=1).value = product_name
        output_sheet.cell(row=row_index, column=2).value = product_id
        output_sheet.cell(row=row_index, column=3).value = price
        row_index += 1
    else:
        # Fuzzy matching to find similar product names
        matched_product, score = process.extractOne(product_name, product_id_map.keys())
        if score > 86:  # Adjust the threshold based on your data
            product_id = product_id_map[matched_product]
            output_sheet.cell(row=row_index, column=1).value = product_name
            output_sheet.cell(row=row_index, column=2).value = product_id
            output_sheet.cell(row=row_index, column=3).value = price
            row_index += 1


# Save the output Excel file
output_file = "Mapped_Products.xlsx"
output_workbook.save(output_file)

# Close all workbooks
winter_top_selling_workbook.close()
overpricing_workbook.close()

print(f"Results have been saved to {output_file}")
